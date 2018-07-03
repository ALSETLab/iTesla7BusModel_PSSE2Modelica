import os
import openpyxl #import the Python module needed to work with the Excel files
import math #import module of mathematical functions
import cmath #import module of mathematical complex functions
import glob



k=1
dir_list = glob.glob("C:\Users\Narasimham\Desktop\Sevenbus\snapshots\N7_*/") # List all directories with PSSE raw files
for w in range(0,len(dir_list)):
    os.chdir(str(dir_list[w])) #set the working directory as the one with Excel files
    #load Excel files downloaded from the Nord Pool website
    '''(there is a problem with the files format,
    they need to be opened and saved as ".xlsx" files'''
    if any(glob.glob("*.raw")):
        for raw in glob.glob("*.raw"):
            os.remove(raw)
    if any(glob.glob("load_distribution.xlsx")):
        os.remove('load_distribution.xlsx')


    k=k+1;
    print('Opening workbooks...')
    ld = openpyxl.Workbook()
    sheet=ld.get_active_sheet()

    print('Creation of the load profile workbook')
    #Row and column headings
    sheet['A1']='Hours'
    sheet['B1']='Bus-1'
    sheet['C1']='Bus-3'
    sheet['D1']='Bus-4'
    sheet['E1']='Bus-5'

    ld_data=openpyxl.load_workbook('data.xlsx')
    sheet1=ld_data.get_active_sheet()


    for i in range(2,26):
        print('Creating load distribution in excel.....')
        sheet.cell(row=i, column=1).value=i-2
        sheet.cell(row=i, column=2).value=0.166667*sheet1.cell(row=i, column=k).value
        sheet.cell(row=i, column=3).value=0.166667*sheet1.cell(row=i, column=k).value
        sheet.cell(row=i, column=4).value=0.333333*sheet1.cell(row=i, column=k).value
        sheet.cell(row=i, column=5).value=0.333333*sheet1.cell(row=i, column=k).value

    print('Running PSSE...')
    import sys
    sys.path.append('C:\Program Files (x86)\PTI\PSSEXplore34\PSSBIN') #make this point to the pssbin directory
    os.environ['PATH']=(r'C:\Program Files (x86)\PTI\PSSEXplore34\PSSBIN;'+os.environ['PATH'])
    import redirect
    redirect.psse2py()
    import psspy
    psspy.throwPsseExceptions = True
    nbuses=50; #max no of buses
    pathname=r"C:\Users\Narasimham\Desktop\Sevenbus\Loads" #path of the folder containing the solved base case
    casefile=os.path.join(pathname,"7Bus_conv.sav") #use the solved base case
    ierr=psspy.psseinit(nbuses)
    psspy.case(casefile)
    for i in range(2,26):
        print('Changing additional loads...')
        psspy.load_chng_4(1, '1', realar1=sheet.cell(row=i, column=2).value, realar2=0.01*sheet.cell(row=i, column=2).value)
        psspy.load_chng_4(3, '1', realar1=sheet.cell(row=i, column=3).value, realar2=0.01*sheet.cell(row=i, column=3).value)
        psspy.load_chng_4(4, '1', realar1=sheet.cell(row=i, column=4).value, realar2=0.01*sheet.cell(row=i, column=4).value)
        psspy.load_chng_4(5, '1', realar1=sheet.cell(row=i, column=5).value, realar2=0.01*sheet.cell(row=i, column=5).value)

        print('Changes completed...')
        psspy.rawd_2(0, 1, [0,0,1,0,0,0,0], 0, "Snap_before_PF.raw") #save the raw file to convert to CIM
        b='h' + str(i-2) + '_before_PF.raw'
        os.rename("Snap_before_PF.raw", b)
        psspy.fnsl([1,0,0,0,0,0,0,0]) #solve the power flow
        ival = psspy.solved() #flag to check power flow convergence
        if ival==0:
            print('Convergence')
            sheet.cell(row=i, column=7).value='Convergence'
            psspy.save('temp.sav') #save temporarily the solved case
            psspy.case('temp.sav') #set the saved case as current case
            psspy.rawd_2(0, 1, [0,0,1,0,0,0,0], 0, "Snap_after_PF.raw") #save the raw file to convert to CIM
            b='h' + str((i-2)) + '_after_PF.raw'
            os.rename("Snap_after_PF.raw", b)
        else:
            print('No convergence')
            sheet.cell(row=i, column=7).value='Non-convergence'

    psspy.close_powerflow()
    #save the Excel file with all data
    ld.save('load_distribution.xlsx')


print('Execution done')
